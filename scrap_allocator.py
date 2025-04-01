import pandas as pd
import re
import sys
import traceback
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
import argparse

print("--- Script Starting ---")

# --- Configuration ---
RECAP_SHEET_NAME = 'By Consumer'
RECAP_STRUCTURE_COL = 'Unnamed: 0'
RECAP_AMOUNT_COL = 'Tons'
WORKSHEET_GRADE_COL = 'Ferrous Product Group & Grade'
WORKSHEET_MILL_COL = 'Mill'
WORKSHEET_TONS_COL = 'Total Available in Sales Month (GT)'

# Define the sheets in the worksheet file that correspond to depots
WORKSHEET_DEPOT_SHEETS = [
    '401Dallas',
    '404 Fort Worth',
    '402 Houston',
    '405 Liberty',
    '407 Bryan',
    '410 Dallas West'
    # Add other sheet names if necessary
]

# --- Mapping Data ---
# Structure: mapping[depot_number_str][worksheet_grade_str] = {'mill': mill_name, 'alias': recap_grade_alias}
# !! IMPORTANT: Review this mapping carefully for accuracy vs. your Excel files !!
mapping = {
    '401': {
        "8BBU - 8B BUSHELING 5'": {'mill': 'Avec (Madil) - LAVE603', 'alias': 'Bush'},
        "8B - 8B (BUSHELING UNPREPARED)": {'mill': 'Avec (Madil) - LAVE603', 'alias': 'Bush'},
        "GM AUTO STAMPING": {'mill': 'Avec (Madil) - LAVE603', 'alias': 'Bush'},
        "PGCS - 3' P&S": {'mill': 'Avec (Madil) - LAVE603', 'alias': 'P&S'},
        "Rail Crop": {'mill': 'Avec (Madil) - LAVE603', 'alias': 'Rail Crops'},
        "Other Rail": {'mill': 'Avec (Madil) - LAVE603', 'alias': 'Rail Crops'},
        "HMS1": {'mill': 'Midlothian - LGER617', 'alias': '#1 HMS'},
        "HMS 1/2 - HMS PREPARED": {'mill': 'Midlothian - LGER617', 'alias': 'HMS'},
        "9A - CAST IRON PREPARED": {'mill': 'Midlothian - LGER617', 'alias': 'Cast'},
        "7B - STEEL TURNINGS": {'mill': 'Midlothian - LGER617', 'alias': 'MST'},
        "Frag Feed (RTIN)": {'mill': 'Midlothian - LGER617', 'alias': 'RTIN'},
        "TINST": {'mill': 'Midlothian - LGER617', 'alias': 'TINST'},
        "FFHMS": {'mill': 'Midlothian - LGER617', 'alias': 'FFHMS'},
        "PUNC - MADIX SLUGS": {'mill': 'HRH Metals', 'alias': 'Slugs'},
        "9BHUB -  FOUNDRY CAST": {'mill': 'Tyler Pipe - LDAV640', 'alias': 'Hubs and Rotors'}
    },
    '404': {
        "Rail Crop": {'mill': 'Avec (Madil) - LAVE603', 'alias': 'Rail Crops'},
        "8BBU - 8B BUSHELING 5'": {'mill': 'Avec (Madil) - LAVE603', 'alias': 'Bush'},
        "8B - 8B (BUSHELING UNPREPARED)": {'mill': 'Avec (Madil) - LAVE603', 'alias': 'Bush'},
        "PUNC - MADIX PUNCHINGS": {'mill': 'Avec (Madil) - LAVE603', 'alias': 'Bush'},
        "PGCS - 3' P&S": {'mill': 'Avec (Madil) - LAVE603', 'alias': 'P&S'},
        "HMS1": {'mill': 'Midlothian - LGER617', 'alias': '#1 HMS'},
        "HMS 1/2 - HMS PREPARED": {'mill': 'Midlothian - LGER617', 'alias': 'HMS'},
        "9A - CAST IRON PREPARED": {'mill': 'Midlothian - LGER617', 'alias': 'Cast'},
        "7B - STEEL TURNINGS": {'mill': 'Midlothian - LGER617', 'alias': 'MST'},
        "Frag Feed (RTIN)": {'mill': 'Midlothian - LGER617', 'alias': 'RTIN'},
        "TINST": {'mill': 'Midlothian - LGER617', 'alias': 'TINST'},
        "FFHMS": {'mill': 'Midlothian - LGER617', 'alias': 'FFHMS'},
        "PUNC - MADIX SLUGS": {'mill': 'HRH Metals', 'alias': 'Slugs'}
    },
    '410': {
        "HMS1": {'mill': 'Midlothian - LGER617', 'alias': '#1 HMS'},
        "HMS 1/2 - HMS PREPARED": {'mill': 'Midlothian - LGER617', 'alias': 'HMS'},
        "9A - CAST IRON PREPARED": {'mill': 'Midlothian - LGER617', 'alias': 'Cast'},
        "7B - STEEL TURNINGS": {'mill': 'Midlothian - LGER617', 'alias': 'MST'},
        "Frag Feed (RTIN)": {'mill': 'Midlothian - LGER617', 'alias': 'RTIN'},
        "TINST": {'mill': 'Midlothian - LGER617', 'alias': 'TINST'},
        "FFHMS": {'mill': 'Midlothian - LGER617', 'alias': 'FFHMS'},
        "8BBU - 8B BUSHELING 5'": {'mill': 'Avec (Madil) - LAVE603', 'alias': 'Bush'},
        "8B - 8B (BUSHELING UNPREPARED)": {'mill': 'Avec (Madil) - LAVE603', 'alias': 'Bush'},
        "GM AUTO STAMPING": {'mill': 'Avec (Madil) - LAVE603', 'alias': 'Bush'},
        "PUNC - MADIX PUNCHINGS": {'mill': 'Avec (Madil) - LAVE603', 'alias': 'Bush'},
        "PGCS - 3' P&S": {'mill': 'Avec (Madil) - LAVE603', 'alias': 'P&S'},
        "Rail Crop": {'mill': 'Avec (Madil) - LAVE603', 'alias': 'Rail Crops'},
        "Other Rail": {'mill': 'Avec (Madil) - LAVE603', 'alias': 'Rail Crops'},
        "PUNC - MADIX SLUGS": {'mill': 'HRH Metals', 'alias': 'Slugs'},
        "9BHUB -  FOUNDRY CAST": {'mill': 'Tyler Pipe - LDAV640', 'alias': 'Hubs and Rotors'}
    },
    '402': {
        "9A - CAST IRON PREPARED": {'mill': 'Midlothian - LGER617', 'alias': 'Cast'},
        "Frag Feed (RTIN)": {'mill': 'Midlothian - LGER617', 'alias': 'RTIN'},
        "TINST": {'mill': 'Midlothian - LGER617', 'alias': 'TINST'},
        "FFHMS": {'mill': 'Midlothian - LGER617', 'alias': 'FFHMS'},
        "9BHUB -  FOUNDRY CAST": {'mill': 'Tyler Pipe - LDAV640', 'alias': 'Hubs and Rotors'},
        "8BBU - 8B BUSHELING 5'": {'mill': 'Optimus -', 'alias': 'Bush'},
        "8B - 8B (BUSHELING UNPREPARED)": {'mill': 'Optimus -', 'alias': 'Bush'},
        "PUNC - MADIX PUNCHINGS": {'mill': 'Optimus -', 'alias': 'Bush'},
        "HMS1": {'mill': 'Optimus -', 'alias': '#1 HMS'},
        "HMS 1/2 - HMS PREPARED": {'mill': 'Optimus -', 'alias': 'HMS'},
        "PGCS - 3' P&S": {'mill': 'Optimus -', 'alias': 'P&S'},
        "Rail Crop": {'mill': 'Optimus -', 'alias': 'Rail Crops'}
    },
    '405': {
        "Frag Feed (RTIN)": {'mill': 'Midlothian - LGER617', 'alias': 'RTIN'},
        "TINST": {'mill': 'Midlothian - LGER617', 'alias': 'TINST'},
        "FFHMS": {'mill': 'Midlothian - LGER617', 'alias': 'FFHMS'},
        "9BHUB -  FOUNDRY CAST": {'mill': 'Tyler Pipe - LDAV640', 'alias': 'Hubs and Rotors'},
        "Rail Crop": {'mill': 'Optimus -', 'alias': 'Rail Crops'},
        "8BBU - 8B BUSHELING 5'": {'mill': 'Optimus -', 'alias': 'Bush'},
        "HMS1": {'mill': 'Optimus -', 'alias': '#1 HMS'},
        "HMS 1/2 - HMS PREPARED": {'mill': 'Optimus -', 'alias': 'HMS'},
        "PGCS - 3' P&S": {'mill': 'Optimus -', 'alias': 'P&S'}
    },
    '407': {
        "7B - STEEL TURNINGS": {'mill': 'Midlothian - LGER617', 'alias': 'MST'},
        "Frag Feed (RTIN)": {'mill': 'Midlothian - LGER617', 'alias': 'RTIN'},
        "TINST": {'mill': 'Midlothian - LGER617', 'alias': 'TINST'},
        "FFHMS": {'mill': 'Midlothian - LGER617', 'alias': 'FFHMS'},
        "9BHUB -  FOUNDRY CAST": {'mill': 'Tyler Pipe - LDAV640', 'alias': 'Hubs and Rotors'},
        "8BBU - 8B BUSHELING 5'": {'mill': 'Jewett - LDAV640', 'alias': 'Bush'},
        "PGCS - 3' P&S": {'mill': 'Jewett - LDAV640', 'alias': 'P&S'},
        "HMS1": {'mill': 'Jewett - LDAV640', 'alias': '#1 HMS'},
        "HMS 1/2 - HMS PREPARED": {'mill': 'Jewett - LDAV640', 'alias': 'HMS'}
    }
}

# --- Helper Functions ---
def get_depot_number(sheet_name):
    """Extracts the first sequence of digits from a sheet name."""
    match = re.search(r'\d+', sheet_name)
    return match.group(0) if match else None

def normalize_grade(grade):
    """Normalizes a grade string for better pattern matching."""
    if not grade:
        return None
    # Remove extra spaces and convert to lowercase
    grade = ' '.join(grade.lower().split())
    # Remove special characters except hyphens
    grade = re.sub(r'[^\w\s-]', '', grade)
    return grade

def find_matching_grade(worksheet_grade, depot_mapping):
    """Finds the best matching grade in the depot mapping."""
    if not worksheet_grade:
        return None
    
    normalized_grade = normalize_grade(worksheet_grade)
    if not normalized_grade:
        return None
    
    # Try exact match first
    if worksheet_grade in depot_mapping:
        return worksheet_grade
    
    # Try normalized match
    for mapping_grade in depot_mapping.keys():
        if normalize_grade(mapping_grade) == normalized_grade:
            return mapping_grade
    
    # Try partial match
    for mapping_grade in depot_mapping.keys():
        if normalized_grade in normalize_grade(mapping_grade) or normalize_grade(mapping_grade) in normalized_grade:
            return mapping_grade
    
    return None

def find_depot_numbers_in_recap_row(text):
    """Finds all depot numbers (like D401, D404) in a string."""
    if not isinstance(text, str):
        return []
    # Look for D followed by digits, ensuring it's not immediately followed by another digit
    # This will match both individual depots (D401) and sections (D401/D404/D410)
    return re.findall(r'D(\d+)(?!\d)', text)

# --- Main Logic ---
def main():
    print("--- Script Starting ---")

    # --- Argument Parsing ---
    parser = argparse.ArgumentParser(description='Process scrap allocation files.')
    parser.add_argument('worksheet_file', help='Path to the input Sales Worksheet Excel file.')
    parser.add_argument('recap_file', help='Path to the input/output Recap Allocation Excel file.')
    args = parser.parse_args()
    print(f"Using Worksheet File: {args.worksheet_file}")
    print(f"Using Recap File: {args.recap_file}")
    # --- End Argument Parsing ---

    # 1. Read Worksheet Data and Aggregate Amounts
    aggregated_amounts = {}
    depot_grand_totals = {} # New dictionary to store total per depot
    print(f"Reading worksheet file: {args.worksheet_file}")
    # Read sheet names first to know which ones exist
    try:
        xls = pd.ExcelFile(args.worksheet_file, engine='openpyxl')
        available_sheets = xls.sheet_names
    except FileNotFoundError:
        print(f"ERROR: Worksheet file not found: {args.worksheet_file}")
        sys.exit(1)
    except Exception as e:
        print(f"ERROR: Could not probe worksheet file sheets: {e}")
        sys.exit(1)

    sheets_to_process = [s for s in WORKSHEET_DEPOT_SHEETS if s in available_sheets]
    if not sheets_to_process:
        print(f"ERROR: None of the configured depot sheets {WORKSHEET_DEPOT_SHEETS} were found in {args.worksheet_file}")
        sys.exit(1)

    for sheet_name in sheets_to_process:
        depot_num = get_depot_number(sheet_name)
        if not depot_num:
            print(f"Warning: Could not extract depot number from sheet name '{sheet_name}'. Skipping.")
            continue

        # -- Set header_index back to fixed Row 3 --
        # header_index = 6 if sheet_name in ['402 Houston', '405 Liberty', '407 Bryan', '410 Dallas West'] else 5 # Dynamic logic
        header_index = 2 # Header is on row 3
        print(f"  Processing sheet: {sheet_name} (Depot {depot_num}), expecting headers in row {header_index + 1}")

        try:
            # Read individual sheet with specific header row (index 2)
            df_sheet = pd.read_excel(args.worksheet_file, sheet_name=sheet_name, engine='openpyxl', header=header_index)
        except Exception as e:
            print(f"  ERROR: Could not read sheet '{sheet_name}'. Error: {e}. Skipping sheet.")
            continue

        # Check if required columns exist by header name
        if WORKSHEET_GRADE_COL not in df_sheet.columns:
            print(f"  ERROR: Grade column '{WORKSHEET_GRADE_COL}' not found in sheet '{sheet_name}'. Skipping sheet.")
            continue
        if WORKSHEET_TONS_COL not in df_sheet.columns:
            print(f"  ERROR: Amount column '{WORKSHEET_TONS_COL}' not found in sheet '{sheet_name}'. Skipping sheet.")
            continue

        for index, row in df_sheet.iterrows():
            # Access columns by header name for reliability
            worksheet_grade_original = str(row[WORKSHEET_GRADE_COL]).strip() if pd.notna(row[WORKSHEET_GRADE_COL]) else None
            tons = pd.to_numeric(row[WORKSHEET_TONS_COL], errors='coerce')

            # Skip if grade is blank/None or amount is zero/NaN
            if not worksheet_grade_original or worksheet_grade_original.isspace() or pd.isna(tons) or tons == 0:
                continue

            depot_mapping = mapping.get(depot_num)
            if not depot_mapping:
                continue

            matching_grade = find_matching_grade(worksheet_grade_original, depot_mapping)

            if not matching_grade:
                print(f"  Warning: Grade '{worksheet_grade_original}' from sheet '{sheet_name}' (Depot {depot_num}) not found in mapping. Skipping.")
                continue
            else:
                grade_info = depot_mapping[matching_grade]
                mill_name = grade_info['mill']
                alias = grade_info['alias']
                key = (depot_num, mill_name, alias)
                aggregated_amounts[key] = aggregated_amounts.get(key, 0) + tons
                # Also add to the depot grand total
                depot_grand_totals[depot_num] = depot_grand_totals.get(depot_num, 0) + tons

    print(f"\nFinished reading worksheet. Aggregated {len(aggregated_amounts)} entries.")
    # Optional: Print depot grand totals for debugging
    # print("\n--- Depot Grand Totals ---")
    # for depot, total in depot_grand_totals.items():
    #     print(f"  Depot {depot}: {total}")
    # print("--------------------------\n")

    # 2. Read Recap Sheet (using fixed header row 6)
    print(f"\nReading recap file: {args.recap_file}, sheet: {RECAP_SHEET_NAME}")
    try:
        # Read using header=5 (Row 6)
        df_recap = pd.read_excel(args.recap_file, sheet_name=RECAP_SHEET_NAME, engine='openpyxl', header=5, keep_default_na=False)

        # Check if needed columns were found using the names from Row 6
        if RECAP_AMOUNT_COL not in df_recap.columns:
            print(f"ERROR: Target amount column '{RECAP_AMOUNT_COL}' (expected C6) not found in header row 6 of sheet '{RECAP_SHEET_NAME}'. Found headers: {list(df_recap.columns)}")
            sys.exit(1)

        # Rename the first column (Column A, likely unnamed or misnamed due to blank A6) to 'Structure'
        first_col_name = df_recap.columns[0]
        print(f"  Renaming first column (originally '{first_col_name}') to '{RECAP_STRUCTURE_COL}' for internal use.")
        df_recap.rename(columns={first_col_name: RECAP_STRUCTURE_COL}, inplace=True)

        # Now check if the renamed structure column exists
        if RECAP_STRUCTURE_COL not in df_recap.columns:
             print(f"ERROR: Failed to access structure column after renaming. Columns: {list(df_recap.columns)}")
             sys.exit(1)

        # Ensure the target amount column is numeric for calculations
        df_recap[RECAP_AMOUNT_COL] = pd.to_numeric(df_recap[RECAP_AMOUNT_COL], errors='coerce').fillna(0)

        # Create a copy for modification
        df_recap_modified = df_recap.copy()
        # IMPORTANT: Clear the target column before populating
        df_recap_modified[RECAP_AMOUNT_COL] = 0.0

    except FileNotFoundError:
        print(f"ERROR: Recap file not found: {args.recap_file}")
        sys.exit(1)
    except ValueError as e: # Handles sheet not found
        print(f"ERROR: Sheet '{RECAP_SHEET_NAME}' not found in {args.recap_file}. Details: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"ERROR: Could not read recap file: {e}")
        sys.exit(1)

    # 3. Process Recap Sheet Rows to Update Amounts
    rows_updated = 0
    current_mill = None
    current_depot_header_text = None
    current_depot_sum = 0
    current_mill_sum = 0
    rows_to_update_with_totals = {}
    rows_to_update_with_grand_totals = {} # New: Store index for Depot Grand Total rows

    print("\nProcessing recap sheet and populating amounts...")
    # First pass: Identify rows, update aliases, and track sums/total row locations
    for index, row in df_recap.iterrows():
        structure_text = str(row[RECAP_STRUCTURE_COL]).strip() if pd.notna(row[RECAP_STRUCTURE_COL]) else ''

        # --- Refined Row Type Identification ---
        is_mill_row = False
        is_depot_header_row = False
        is_alias_row = False
        is_total_row = False
        is_grand_total_row = False # New flag
        total_row_type = None

        known_mills = set(info['mill'] for depot_map in mapping.values() for info in depot_map.values())
        known_aliases = set(info['alias'] for depot_map in mapping.values() for info in depot_map.values())

        # --- Hardcoded Skip for Specific Unmapped Mills ---
        if structure_text in ['CMC - LCMC606', 'East Jordan - LEJO601']:
            # print(f"DEBUG: Hardcoded skip - Resetting mill context for: '{structure_text}'")
            current_mill = None # Ensure subsequent lookups fail
            # Potentially clear sums if needed, but likely handled by 0 lookup
            continue # Skip further processing for this specific header row
        # --- End Hardcoded Skip ---

        if structure_text in known_mills:
            is_mill_row = True
        else:
            # Determine row type if not a known mill or explicitly skipped header
            depots_in_row = find_depot_numbers_in_recap_row(structure_text)
            # Check for Depot Grand Total first (e.g., "Total GT D401")
            grand_total_match = re.match(r"Total GT D(\d+)", structure_text)
            if grand_total_match:
                is_grand_total_row = True
                grand_total_depot = grand_total_match.group(1)

            elif depots_in_row and structure_text not in known_aliases and not any(alias in structure_text for alias in known_aliases if len(alias)>2):
                 if structure_text.startswith('D') or ' - D' in structure_text:
                    is_depot_header_row = True

            # Check for other Total rows (Mill or standard Depot total)
            elif structure_text.startswith('Total'):
                 is_total_row = True
                 if current_mill and current_mill.split(' - ')[0] in structure_text:
                     total_row_type = 'mill'
                 else:
                     total_row_type = 'depot'

            elif structure_text and current_mill:
                 is_alias_row = True
        # --- End Refined Row Type Identification ---

        if is_mill_row:
            current_mill = structure_text
            current_depot_header_text = None
            current_depot_sum = 0
            current_mill_sum = 0
            continue

        if not current_mill and not is_grand_total_row: # Allow Grand Totals even without mill context
            continue

        if is_depot_header_row:
            current_depot_header_text = structure_text
            current_depot_sum = 0
            continue

        if is_grand_total_row:
            # Store index and the depot number it applies to
            rows_to_update_with_grand_totals[index] = grand_total_depot
            continue

        if is_total_row:
             sum_to_store = current_mill_sum if total_row_type == 'mill' else current_depot_sum
             rows_to_update_with_totals[index] = {'type': total_row_type, 'value': sum_to_store}
             continue

        # Process as Grade Alias row
        if is_alias_row and current_depot_header_text:
            recap_alias_raw = structure_text
            depots_for_this_row = find_depot_numbers_in_recap_row(current_depot_header_text)

            # --- Alias Extraction Logic (keep) ---
            if recap_alias_raw in known_aliases:
                recap_alias_lookup = recap_alias_raw
            else:
                base_alias_found = None
                sorted_known_aliases = sorted(list(known_aliases), key=len, reverse=True)
                for known_alias in sorted_known_aliases:
                    if recap_alias_raw.startswith(known_alias) and \
                       (len(recap_alias_raw) == len(known_alias) or recap_alias_raw[len(known_alias)] in [' ', '-']):
                        base_alias_found = known_alias
                        break
                recap_alias_lookup = base_alias_found if base_alias_found else recap_alias_raw
            # --- End Alias Extraction Logic ---

            if not depots_for_this_row or not recap_alias_lookup:
                continue

            total_amount_for_row = 0
            found_match_for_row = False
            for depot_num in depots_for_this_row:
                key = (depot_num, current_mill, recap_alias_lookup)
                amount = aggregated_amounts.get(key, 0)
                if amount != 0:
                    total_amount_for_row += amount
                    found_match_for_row = True

            if found_match_for_row:
                # -- REMOVE DEBUG: Print context JUST BEFORE writing --
                # print(f"DEBUG: Writing Alias - Row: {index + 7}, Structure: '{structure_text}', MillContext: '{current_mill}', DepotHeader: '{current_depot_header_text}', AliasLookup: '{recap_alias_lookup}', Amount: {total_amount_for_row}")
                # -- END REMOVE --
                if RECAP_AMOUNT_COL in df_recap_modified.columns:
                    if pd.api.types.is_number(total_amount_for_row):
                        df_recap_modified.loc[index, RECAP_AMOUNT_COL] = total_amount_for_row
                        current_depot_sum += total_amount_for_row
                        current_mill_sum += total_amount_for_row
                        rows_updated += 1
                else:
                    print(f"    -> ERROR: Target amount column '{RECAP_AMOUNT_COL}' not found.")
            else:
                 pass

        elif is_alias_row and not current_depot_header_text:
             pass

    # --- Second Pass: Update the identified Total Rows ---
    print("\nUpdating Mill/Depot Total Rows...")
    for index, total_info in rows_to_update_with_totals.items():
        total_value = total_info['value']
        total_type = total_info['type']
        if pd.api.types.is_number(total_value):
             if RECAP_AMOUNT_COL in df_recap_modified.columns:
                 df_recap_modified.loc[index, RECAP_AMOUNT_COL] = total_value
                 rows_updated += 1
             else:
                  print(f"  -> ERROR: Cannot update {total_type.upper()} Total Row {index+7}. Target column '{RECAP_AMOUNT_COL}' not found.")
        else:
            print(f"  -> WARNING: Skipping {total_type.upper()} Total Row {index+7}. Calculated sum '{total_value}' is not numeric.")

    # --- Third Pass: Update the Depot Grand Total Rows ---
    print("\nUpdating Depot Grand Total Rows...")
    for index, depot_num in rows_to_update_with_grand_totals.items():
        grand_total_value = depot_grand_totals.get(depot_num, 0) # Get the pre-calculated total
        if pd.api.types.is_number(grand_total_value):
             if RECAP_AMOUNT_COL in df_recap_modified.columns:
                 df_recap_modified.loc[index, RECAP_AMOUNT_COL] = grand_total_value
                 rows_updated += 1
             else:
                  print(f"  -> ERROR: Cannot update Grand Total Row {index+7} for Depot {depot_num}. Target column '{RECAP_AMOUNT_COL}' not found.")
        else:
             print(f"  -> WARNING: Skipping Grand Total Row {index+7} for Depot {depot_num}. Calculated sum '{grand_total_value}' is not numeric.")
    # --- End Updates ---

    print(f"\nFinished processing recap sheet. Updated {rows_updated} rows (including totals).")

    # 4. Save Updated Recap File
    print(f"Saving updated data back to {args.recap_file}, sheet: {RECAP_SHEET_NAME}...")
    try:
        # Load the existing workbook, ensure formulas are read (data_only=False)
        wb = load_workbook(args.recap_file, data_only=False)
        ws = wb[RECAP_SHEET_NAME]

        print("  Checking cells and updating non-formula cells only...")
        cells_updated_values = 0
        cells_skipped_formulas = 0

        # Update only the values in the Tons column (Column C), skipping formulas
        for index, row_data in df_recap_modified.iterrows():
            # Add 7 because Excel is 1-indexed and we have a header row at row 6
            target_row_excel = index + 7
            target_col_excel = 3 # Column C

            # Get the cell object
            target_cell = ws.cell(row=target_row_excel, column=target_col_excel)

            # Check if the cell contains a formula
            if target_cell.data_type == 'f':
                # print(f"    Skipping cell {target_cell.coordinate} - contains formula.") # Optional debug print
                cells_skipped_formulas += 1
            else:
                # Cell doesn't contain a formula, update its value
                calculated_value = row_data[RECAP_AMOUNT_COL]
                # Only write if the value needs changing
                if target_cell.value != calculated_value:
                     # -- REMOVE DEBUG: Print Save Action --
                     # print(f"DEBUG: Saving - Cell: {target_cell.coordinate}, OldValue: {target_cell.value}, NewValue: {calculated_value}")
                     # -- END REMOVE --
                     target_cell.value = calculated_value
                     cells_updated_values += 1
                 # else: # If value is already correct, don't count as update

        print(f"  Finished checking: Updated {cells_updated_values} non-formula cells, skipped {cells_skipped_formulas} formula cells.")

        # Save while preserving formatting
        wb.save(args.recap_file)
        print("File saved successfully.")
    except PermissionError:
        print(f"\nERROR: Permission denied. Could not save '{args.recap_file}'.")
        print("Please ensure the file is closed in Excel and you have write permissions.")
        sys.exit(1)
    except Exception as e:
        print(f"\nERROR: Could not save the updated recap file: {e}")
        traceback.print_exc() # Print full traceback for saving errors
        sys.exit(1)

    print("\nScript finished.")

# --- Run Script --- (Ensure this is the VERY end of the file)
if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\nAn unexpected error occurred during script execution:")
        print(f"ERROR TYPE: {type(e).__name__}")
        print(f"ERROR DETAILS: {e}")
        print("--- Full Traceback ---")
        traceback.print_exc()
        sys.exit(1) # Exit with error code if main fails 